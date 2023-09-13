import csv
import requests
import time
import yagmail
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import schedule
import threading

# Tu clave API de IQAir
API_KEY = "API_KEY_IQAir"

# Función para obtener recomendaciones de salud en función del AQI
def get_health_recommendations(aqi):
    if aqi <= 50:
        return "Calidad del aire satisfactoria. Riesgo mínimo."
    elif 51 <= aqi <= 100:
        return "Calidad del aire aceptable. Sin riesgo para la mayoría, pero podría haber riesgos para ciertas personas muy sensibles."
    elif 101 <= aqi <= 150:
        return "Puede haber riesgos para la salud de ciertos grupos sensibles, como niños menores de 5 años."
    elif 151 <= aqi <= 200:
        return "Riesgos para la salud de todos, más severos para grupos de riesgo. No apto para niños menores de 5 años."
    elif 201 <= aqi <= 300:
        return "Advertencia de salud: todos pueden experimentar efectos de salud más graves. No apto para niños menores de 5 años."
    elif aqi > 300:
        return "Advertencia de salud: toda la población puede experimentar efectos de salud más graves. No apto para niños menores de 5 años."
    else:
        return "N/A"

# Función para obtener datos de calidad del aire
def fetch_air_quality_data(lat, lon):
    API_ENDPOINT = f"http://api.airvisual.com/v2/nearest_city?lat={lat}&lon={lon}&key={API_KEY}"

    try:
        response = requests.get(API_ENDPOINT)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Falló al obtener datos de calidad del aire: {response.content}")
            return None
    except Exception as e:
        print(f"Ocurrió un error: {e}")
        return None

# Función para aplicar métricas AQI a una columna
def apply_metrics(aqi_value):
    if 0 <= aqi_value <= 50:
        return "Verde"
    elif 51 <= aqi_value <= 100:
        return "Amarillo"
    elif 101 <= aqi_value <= 150:
        return "Naranja"
    elif 151 <= aqi_value <= 200:
        return "Rojo"
    elif 201 <= aqi_value <= 300:
        return "Morado"
    else:
        return "Granate"

# Función para generar y enviar el informe
def send_report():
    # Leer archivo de ubicaciones
    with open('locations.csv', mode='r') as infile:
        reader = csv.reader(infile)
        next(reader, None)  # Saltar fila de encabezado

        # Escribir archivo de historial
        with open('history.csv', mode='w', newline='') as outfile:
            fieldnames = ['Localidad', 'Latitud', 'Longitud', 'AQI_US', 'Contaminante Principal US', 'Recomendaciones']
            writer = csv.DictWriter(outfile, fieldnames=fieldnames)

            writer.writeheader()

            for row in reader:
                if len(row) == 3:
                    localidad, lat, lon = row
                elif len(row) == 2:
                    lat, lon = row
                    localidad = "N/A"
                else:
                    print(f"Saltando fila inválida: {row}")
                    continue

                time.sleep(5)  # Añadir un retardo de 5 segundos entre cada localidad

                data = fetch_air_quality_data(float(lat), float(lon))
                if data:
                    current_data = data.get('data', {}).get('current', {})
                    pollution_data = current_data.get('pollution', {})
                    aqi_us = pollution_data.get('aqius', 'N/A')
                    main_us = pollution_data.get('mainus', 'N/A')

                    recommendations = get_health_recommendations(aqi_us)

                    writer.writerow({
                        'Localidad': localidad,
                        'Latitud': lat,
                        'Longitud': lon,
                        'AQI_US': aqi_us,
                        'Contaminante Principal US': main_us,
                        'Recomendaciones': recommendations
                    })

                time.sleep(10)  # Añadir un retardo de 10 segundos entre cada llamada a la API

    # Leer el archivo CSV original
    file_path = 'history.csv'
    df = pd.read_csv(file_path, encoding='ISO-8859-1')  # especificar codificación

    # Guardar el DataFrame en un archivo Excel (.xlsx)
    excel_file_path = 'modified_history.xlsx'
    df.to_excel(excel_file_path, index=False)

    # Abrir el archivo Excel para editarlo
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    # Iterar sobre las filas para aplicar colores de fondo a la columna AQI_US
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        for cell in row:
            aqi_value = cell.value
            color_flag = apply_metrics(aqi_value)
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            if color_flag == "Verde":
                fill.start_color.rgb = "00FF00"
                fill.end_color.rgb = "00FF00"
            elif color_flag == "Amarillo":
                fill.start_color.rgb = "FFFF00"
                fill.end_color.rgb = "FFFF00"
            elif color_flag == "Naranja":
                fill.start_color.rgb = "FFA500"
                fill.end_color.rgb = "FFA500"
            elif color_flag == "Rojo":
                fill.start_color.rgb = "FF0000"
                fill.end_color.rgb = "FF0000"
            elif color_flag == "Morado":
                fill.start_color.rgb = "800080"
                fill.end_color.rgb = "800080"
            elif color_flag == "Granate":
                fill.start_color.rgb = "800000"
                fill.end_color.rgb = "800000"
            cell.fill = fill

    # Guardar los cambios en el archivo Excel
    wb.save(excel_file_path)

    # Configurar las credenciales de la cuenta de Gmail
    email_address = "your_email@gmail.com"
    password = "your_password"

    # Crear un cliente de yagmail
    yag = yagmail.SMTP(email_address, password)

    # Destinatario y asunto del correo electrónico
    to = "recipient@example.com"

    # Obtener la fecha actual y formatearla como parte del asunto
    current_date = datetime.now().strftime("%Y-%m-%d")
    subject = f"Informe de Calidad del Aire con Métricas de la EPA - {current_date}"

    # URL de la imagen de encabezado
    header_image_url = "URL_IMAGEN.png"

    # Cuerpo del mensaje HTML con imagen de encabezado desde una URL
    message = f"""
    <html>
    <head></head>
    <body>
    <img src="{header_image_url}" alt="Encabezado" width="100%"> <!-- Ajusta al ancho de la página -->
    <h1 style="font-size: 24px;">Informe de Calidad del Aire para Menores de 5 Años</h1>
    <em>Un análisis basado en el Índice de Calidad del Aire de la EPA de EE. UU.</em>
    <p><strong>Este es un mensaje automático. Por favor, no responda a este correo electrónico.</strong></p>
    <p>Estimado Nombre y Puesto</p>
    <p>Es un honor presentarle nuestro Informe de Calidad del Aire para Menores de 5 Años, el cual se basa en el Índice de Calidad del Aire (AQI) de la Agencia de Protección Ambiental de los Estados Unidos (EPA). Nuestra organización está comprometida con la salud y el bienestar de los niños, y este informe tiene como objetivo proporcionarle información crítica sobre la calidad del aire en nuestra localidad.</p>
    <h2 style="font-size: 20px; margin-top: 16px;">Recomendaciones de la EPA:</h2>
    <p>Para obtener detalles adicionales sobre las métricas y las recomendaciones específicas de la EPA, le invitamos a consultar el siguiente enlace: <a href="https://www.epa.gov/air-research">Fuentes y Métricas de Calidad del Aire de la EPA</a></p>
    <h2 style="font-size: 20px; margin-top: 16px;">Fase de Prueba:</h2>
    <p>Este informe es parte de nuestra fase de prueba, donde utilizamos los datos de calidad del aire proporcionados por los sensores de AirVisual más cercanos a las coordenadas geográficas de nuestra localidad. Esto nos permite evaluar y mejorar constantemente la calidad de nuestros informes.</p>
    <h2 style="font-size: 20px; margin-top: 16px;">Leyenda del Informe:</h2>
    <ul>
    <li><strong style="font-weight: bold; color: #00FF00;">Bandera Verde (0 a 50 - Bueno):</strong> La calidad del aire se considera satisfactoria y la contaminación atmosférica presenta un riesgo escaso o nulo.</li>
    <li><strong style="font-weight: bold; color: #FFFF00;">Bandera Amarillo (51 a 100 - Moderado):</strong> La calidad del aire es aceptable, pero podría existir una preocupación moderada para la salud de personas excepcionalmente sensibles.</li>
    <li><strong style="font-weight: bold; color: #FFA500;">Bandera Naranja (101 a 150 - Insalubre para grupos sensibles):</strong> Los miembros de grupos sensibles pueden padecer efectos en la salud. No afectará a las personas en general.</li>
    <li><strong style="font-weight: bold; color: #FF0000;">Bandera Roja (151 a 200 - Insalubres):</strong> Todos pueden comenzar a padecer efectos en la salud y los miembros de grupos sensibles pueden padecer efectos más graves.</li>
    <li><strong style="font-weight: bold; color: #800080;">Bandera Morada (201 a 300 - Muy insalubre):</strong> Advertencias sanitarias de condiciones de emergencia. Mayor probabilidad de que toda la población esté afectada.</li>
    <li><strong style="font-weight: bold; color: #8B0000;">Bandera Granate (301 y superior - Peligroso):</strong> Alerta sanitaria: todos pueden padecer efectos sanitarios más graves.</li>
    </ul>
    <p>El informe completo se encuentra adjunto en el archivo "history.csv" para su revisión. Si tiene alguna pregunta o necesita asistencia adicional, no dude en ponerse en contacto con nuestro equipo de soporte técnico en la dirección de correo electrónico sistemas@fdhp.org.</p>
    <p>Agradecemos su compromiso con la seguridad y el bienestar de nuestros niños.</p>
    </body>
    </html>
    """

    # Adjuntar el archivo Excel modificado
    archivo_adjunto = excel_file_path

    # Enviar el correo electrónico con la imagen de encabezado desde una URL
    yag.send(
        to=to,
        subject=subject,
        contents=message,
        attachments=archivo_adjunto
    )

    # Cerrar la conexión SMTP
    yag.close()

# Programar la ejecución del informe a las 7:00 a.m. y 12:00 p.m. todos los días
schedule.every().day.at("07:00").do(send_report)
schedule.every().day.at("12:00").do(send_report)

# Función para ejecutar el planificador de tareas en segundo plano
def run_schedule():
    while True:
        schedule.run_pending()
        time.sleep(1)

# Iniciar un hilo para ejecutar el planificador de tareas
if __name__ == "__main__":
    scheduler_thread = threading.Thread(target=run_schedule)
    scheduler_thread.start()
